from openpyxl import load_workbook
def Get_data( a,b):

    workbook = load_workbook(r'F:\get_data.xlsx')  # 获取本地文件
    workbooksheet = workbook.active  # 获取文件sheet页
    row = workbooksheet.rows  # 获取行
    column = workbooksheet.columns  # 获取列
    content = workbooksheet.cell(row = a,column = b).value  # 获取参数内容
    return content